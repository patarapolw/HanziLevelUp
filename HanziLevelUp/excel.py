import openpyxl as px
import regex

from CJKhyperradicals.sentence import SpoonFed
from CJKhyperradicals.dict import HanziDict, Cedict
from HanziLevelUp.hanzi import get_all_hanzi, HanziLevel
from HanziLevelUp.vocab import get_all_vocab_plus, VocabToSentence
from webapp.databases import Sentence

HANZI_HEADER = ['Hanzi', 'Pinyin', 'English', 'Heisig', 'Variant', 'Kanji',
                'Level', 'Note', 'Tags']
VOCAB_HEADER = ['Entry', 'Simplified', 'Traditional', 'Pinyin', 'English', 'Sentence',
                'Level', 'Note', 'Tags']
SENTENCE_HEADER = ['Sentence', 'Pinyin', 'English',
                   'Level', 'Note', 'Tags']


class ExcelExport:
    def __init__(self, filename: str):
        self.filename = filename
        self.hanzi_formatter = HanziFormatter()
        self.vocab_formatter = VocabFormatter()
        self.sentence_formatter = SentenceFormatter()

        try:
            self.workbook = px.load_workbook(self.filename)
            if not self.verify():
                print('Invalid file.')
                self.filename += '_'
                self.workbook = self.create()
        except FileNotFoundError:
            self.workbook = self.create()

        self.pre_existing = {
            'Hanzi': set([cell.value for cell in next(self.workbook['Hanzi'].iter_cols())[1:]]),
            'Vocab': set([cell.value for cell in next(self.workbook['Vocab'].iter_cols())[1:]]),
            'Sentence': set([cell.value for cell in next(self.workbook['Sentence'].iter_cols())[1:]])
        }

    @staticmethod
    def create():
        wb = px.Workbook()

        wb.create_sheet('Hanzi')
        ws = wb['Hanzi']
        ws.append(HANZI_HEADER)

        wb.create_sheet('Vocab')
        ws = wb['Vocab']
        ws.append(VOCAB_HEADER)

        wb.create_sheet('Sentence')
        ws = wb['Sentence']
        ws.append(SENTENCE_HEADER)

        return wb

    def verify(self):
        if not all([sheet in self.workbook.sheetnames for sheet in ['Hanzi', 'Vocab', 'Sentence']]):
            return False

        first_row = [cell.value for cell in next(self.workbook['Hanzi'].iter_rows())]
        if first_row != HANZI_HEADER:
            return False

        first_row = [cell.value for cell in next(self.workbook['Vocab'].iter_rows())]
        if first_row != VOCAB_HEADER:
            return False

        first_row = [cell.value for cell in next(self.workbook['Sentence'].iter_rows())]
        if first_row != SENTENCE_HEADER:
            return False

        return True

    def from_db(self):
        ws = self.workbook['Hanzi']
        for hanzi in get_all_hanzi():
            if hanzi not in self.pre_existing['Hanzi']:
                ws.append(self.hanzi_formatter.format(hanzi))

        ws = self.workbook['Vocab']
        for _, vocab in get_all_vocab_plus():
            if vocab not in self.pre_existing['Vocab']:
                ws.append(self.vocab_formatter.format(vocab))

        ws = self.workbook['Sentence']
        for sent_query in Sentence.query:
            sentence = sent_query.sentence
            if sentence not in self.pre_existing['Sentence']:
                ws.append(self.sentence_formatter.format(sentence))

        self.save()

    def save(self):
        self.workbook.save(self.filename)


class HanziFormatter:
    def __init__(self):
        self.hanzi_dict = HanziDict()
        self.hanzi_level = HanziLevel()

    def format(self, hanzi):
        print(hanzi)

        entry = self.hanzi_dict.entries.get(hanzi, dict())
        result = [
            hanzi,
            entry.get('Pin1Yin1', ''),
            entry.get('Meaning', ''),
            entry.get('Heisig', ''),
            entry.get('Variant', ''),
            entry.get('Kanji', ''),
            self.hanzi_level.get_hanzi_level(hanzi),
            '',
            ''
        ]
        assert len(result) == len(HANZI_HEADER), 'Invalid HanziFormatter'

        return result


class VocabFormatter:
    def __init__(self):
        self.cedict = Cedict()
        self.hanzi_level = HanziLevel()
        self.vocab_to_sentence = VocabToSentence()

    def format(self, entry):
        print(entry)

        dict_result = list(self.cedict.search_vocab(entry))
        simplified = ', '.join([item[1] for item in dict_result])
        traditional = ', '.join([item[0] for item in dict_result])
        pinyin = ', '.join([item[2] for item in dict_result])
        english = ', '.join([item[3] for item in dict_result])
        sentences = self.vocab_to_sentence.convert(entry, online=False)
        # traditional = pinyin = english = ''

        level = max([self.hanzi_level.get_hanzi_level(hanzi) for hanzi in entry])

        result = [
            entry,
            simplified,
            traditional,
            pinyin,
            english,
            '\n'.join(['\n'.join(pair) for pair in sentences]),
            level,
            '',
            ''
        ]
        assert len(result) == len(VOCAB_HEADER), 'Invalid VocabFormatter'

        return result


class SentenceFormatter:
    def __init__(self):
        self.spoon_fed = SpoonFed()
        self.hanzi_level = HanziLevel()

    def format(self, sentence):
        print(sentence)

        lookup = list(self.spoon_fed.get_sentence(sentence))
        if len(lookup) > 0:
            english = lookup[0][1]
            try:
                pinyin = lookup[0][2]
            except IndexError:
                pinyin = ''
        else:
            english = pinyin = ''

        result = [
            sentence,
            pinyin,
            english,
            ', '.join([str(self.hanzi_level.get_hanzi_level(char)) for char in sentence
                       if regex.match(r'\p{IsHan}', char)]),
            '',
            ''
        ]
        assert len(result) == len(SENTENCE_HEADER), 'Invalid SentenceFormatter'

        return result
